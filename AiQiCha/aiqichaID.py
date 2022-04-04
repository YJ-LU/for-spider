import time
import random

from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook


def getUrls():
    wb = load_workbook("/Users/apple/Desktop/上市后备企业名单.xlsx")

    ws = wb.active
    ws_col = ws['A']
    url_list = []
    for cell in ws_col:
        url = "https://aiqicha.baidu.com/s?q=" + cell.value + "&t=0"
        url_list.append(url)
    print(url_list)
    return url_list


def main():
    PATH = "/Users/apple/Desktop/code/chromedriver"
    driver = webdriver.Chrome(PATH)

    # name = "武汉数信科技有限公司"
    # url = "https://aiqicha.baidu.com/s?q=" + name + "&t=0"
    list_one = []
    list_two = []
    urls = getUrls()
    for url in urls:
        driver.get(url)

        html = driver.page_source
        h3 = BeautifulSoup(html, "html.parser").find_all("h3", class_='title')
        for r in h3:
            a = r.find("a")
            try:
                next_url = "https://aiqicha.baidu.com" + a["href"]
                list_one.append(next_url)
                print(next_url)
            except TypeError:
                pass
        list_two.append(list_one[0])
        print("list_two is:", list_two)
        list_one.clear()
        time.sleep(random.randint(5, 10))

    driver.close()


if __name__ == '__main__':
    main()
